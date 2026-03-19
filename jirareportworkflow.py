import os
import sys
import time
import logging
import shutil
import pandas as pd
import openpyxl
import warnings
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import io

warnings.filterwarnings('ignore', message='Could not infer format, so each element will be parsed individually')
warnings.filterwarnings('ignore', category=UserWarning, module='pandas')

try:
    import win32com.client as win32
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False


# ========================= CONFIGURATION =========================
class IntegratedJiraConfig:
    """Configuration for the complete Jira automation system."""
    def __init__(self):
        # Jira Authentication Configuration
        self.jira_username = os.getenv('JIRA_USERNAME', 'admin.sbisht@aboutsib.com')
        self.jira_password = os.getenv('JIRA_PASSWORD', 'xxxxxxxxxxxxx')  # <- Fill your Jira password or API token

        self.jira_base_url = "https://aboutsib.atlassian.net"
        self.jira_login_url = (
            "https://id.atlassian.com/login"
            "?continue=https%3A%2F%2Fid.atlassian.com%2Fjoin%2Fuser-access%3Fresource%3Dari%253Acloud%253Ajira%253A%253Asite%252F"
            "f9da0739-6a64-4d76-a380-c07c2f64e2c6%26continue%3Dhttps%253A%252F%252Faboutsib.atlassian.net%252Fjira"
            "&application=jira"
        )
        self.jira_filter_url = f"{self.jira_base_url}/issues/?filter=-4"

        # Date Range Configuration - split into ranges
        self.date_ranges = [
            {"start": "2025-01-01", "end": "2025-04-01", "label": "Q1", "end_operator": "<"},
            {"start": "2025-04-01", "end": "2025-08-01", "label": "Q2-Q3", "end_operator": "<"},
            {"start": "2025-08-01", "end": None, "label": "Q3-Q4", "end_operator": None}
        ]

        # Base JQL Query (without date filter)
        self.base_jql = 'project IN (HELP, VOFB)'

        # File Paths Configuration - Dynamic based on script/executable location
        if getattr(sys, 'frozen', False):
            # Running as compiled executable
            base_path = os.path.dirname(sys.executable)
        else:
            # Running as Python script
            base_path = os.path.dirname(os.path.abspath(__file__))

        self.download_path = os.path.join(base_path, "Downloads")
        self.template_path = os.path.join(base_path, "JiraReportTemplate.xlsx")
        self.output_folder = os.path.join(base_path, "Reports")

        # Create directories if they don't exist
        os.makedirs(self.download_path, exist_ok=True)
        os.makedirs(self.output_folder, exist_ok=True)

        # Email Configuration 
        self.send_email = True
        self.email_via_outlook = True  
        self.email_config = {
            'fromaddr': "siddhant.bisht@aboutsib.com",
            'toaddr': ['indiait@aboutsib.com'],
            'cc': ['john.k@vector97.com'],
            'bcc': ['john.k@vector97.com'],
        }

        # Failure email configuration (Outlook)
        self.failure_email_config = {
            'fromaddr': "siddhant.bisht@aboutsib.com",
            'toaddr': ['siddhant.bisht@aboutsib.com'],
            'cc': [],
            'bcc': [],
        }

        self.email_sheet = "Progress Tracker"

        # Browser and Timeout Settings
        self.default_timeout = 30000
        self.login_timeout = 45000
        self.export_timeout = 180000  # 3 minutes for large exports
        self.headless_mode = False  # Set True to run without UI


# ========================= MAIN AUTOMATION CLASS =========================
class IntegratedJiraAutomation:
    """Complete Jira automation system - download multiple ranges, process, Excel, and email via Outlook."""

    def __init__(self, config: IntegratedJiraConfig):
        self.config = config
        self.logger = self._setup_logging()
        self.browser = None
        self.context = None
        self.page = None
        self.downloaded_csv_paths = []

    def _setup_logging(self) -> logging.Logger:
        """Set up comprehensive logging."""
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)

        if not logger.handlers:
            formatter = logging.Formatter(
                '%(asctime)s - %(levelname)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )

            console_handler = logging.StreamHandler(sys.stdout)
            console_handler.setLevel(logging.INFO)
            console_handler.setFormatter(formatter)

            # Create a timestamped log file in the working directory
            log_filename = f'jira_automation_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
            file_handler = logging.FileHandler(log_filename, encoding='utf-8')
            file_handler.setLevel(logging.DEBUG)
            file_handler.setFormatter(formatter)

            logger.addHandler(console_handler)
            logger.addHandler(file_handler)
            logger.propagate = False

        return logger

    # ========================= JIRA DOWNLOAD AUTOMATION =========================

    def _setup_browser_context(self):
        """Initialize browser and context with download configuration."""
        try:
            download_path = Path(self.config.download_path)
            download_path.mkdir(parents=True, exist_ok=True)
            self.logger.info(f"[SETUP] Download directory: {download_path.absolute()}")

            playwright = sync_playwright().start()

            # Use Chromium for portability (Edge channel optional)
            self.browser = playwright.chromium.launch(
                headless=self.config.headless_mode,
                args=[
                    "--start-maximized",
                    "--disable-dev-shm-usage",
                    "--disable-extensions",
                    "--no-sandbox",
                    "--disable-gpu",
                    "--disable-features=VizDisplayCompositor",
                    "--disable-web-security",
                    "--force-device-scale-factor=1"
                ]
            )

            self.context = self.browser.new_context(
                accept_downloads=True,
                no_viewport=True
            )

            self.page = self.context.new_page()
            self.page.set_default_timeout(self.config.default_timeout)

            try:
                # Maximize window via JS (best effort)
                self.page.evaluate("""
                    () => {
                        if (window.screen && window.screen.width && window.screen.height) {
                            window.resizeTo(window.screen.width, window.screen.height);
                            window.moveTo(0, 0);
                        }
                    }
                """)
                self.logger.info("[DISPLAY] Browser window maximized to full screen")
            except Exception as display_error:
                self.logger.warning(f"[WARNING] Could not maximize window: {display_error}")

            self.logger.info("[SUCCESS] Browser initialized successfully")

        except Exception as e:
            self.logger.error(f"[ERROR] Failed to initialize browser: {e}")
            raise Exception(f"Browser initialization failed: {e}")

    def _wait_and_interact(self, selector_type: str, selector, action: str,
                           value: str = None, timeout: int = None):
        """Generic method to wait for elements and interact with them."""
        if timeout is None:
            timeout = self.config.default_timeout

        try:
            if selector_type == 'test_id':
                element = self.page.get_by_test_id(selector)
            elif selector_type == 'role':
                role, name = selector
                element = self.page.get_by_role(role, name=name)
            elif selector_type == 'text':
                element = self.page.get_by_text(selector)
            elif selector_type == 'css':
                element = self.page.locator(selector)
            else:
                raise ValueError(f"Unknown selector type: {selector_type}")

            element.wait_for(state="visible", timeout=timeout)

            if action == 'click':
                element.click()
                self.logger.debug(f"[CLICK] Clicked: {selector}")
            elif action == 'fill':
                element.click()
                element.fill(value)
                self.logger.debug(f"[FILL] Filled: {selector}")

        except PlaywrightTimeoutError:
            self.logger.error(f"[TIMEOUT] Element not found: {selector}")
            raise Exception(f"Element '{selector}' not found within {timeout}ms")
        except Exception as e:
            self.logger.error(f"[ERROR] Error with {selector}: {e}")
            raise Exception(f"Failed to {action} element '{selector}': {e}")

    def dismiss_2fa_popup(self):
        """Dismiss the two-step verification reminder popup if it appears."""
        try:
            self.logger.info("[2FA] Checking for two-step verification popup...")
            time.sleep(1)

            dismiss_selectors = [
                "text=/remind me later/i",
                "text=/not now/i",
                "text=/skip/i",
                "text=/maybe later/i",
                "button:has-text('Remind me later')",
                "button:has-text('Not now')",
                "button:has-text('Skip')",
                "[data-testid*='remind']",
                "[data-testid*='skip']",
                "[data-testid*='later']"
            ]

            for selector in dismiss_selectors:
                try:
                    dismiss_button = self.page.locator(selector).first
                    if dismiss_button.is_visible(timeout=2000):
                        dismiss_button.click()
                        self.logger.info(f"[SUCCESS] Dismissed 2FA popup using selector: {selector}")
                        time.sleep(1)
                        break
                except Exception:
                    continue

            self.logger.info("[INFO] 2FA popup not present or already dismissed")
            return True

        except Exception as e:
            self.logger.warning(f"[WARNING] Error checking for 2FA popup: {e}")
            return True

    def login_to_jira(self):
        """Handle the complete Jira login process."""
        try:
            self.logger.info("[LOGIN] Starting login process")

            self.page.goto(self.config.jira_login_url, timeout=self.config.login_timeout)

            self.logger.info("[INPUT] Entering username")
            self._wait_and_interact('test_id', 'username', 'fill', self.config.jira_username)
            time.sleep(2)

            self.logger.info("[CLICK] Clicking continue")
            self._wait_and_interact('test_id', 'login-submit-idf-testid', 'click')
            time.sleep(3)

            self.logger.info("[INPUT] Entering password")
            self._wait_and_interact('test_id', 'password', 'fill', self.config.jira_password)
            time.sleep(2)

            self.logger.info("[SUBMIT] Submitting login")
            self._wait_and_interact('test_id', 'login-submit-idf-testid', 'click')

            self.page.wait_for_url("**/jira/**", timeout=self.config.login_timeout)
            self.logger.info("[SUCCESS] Login successful")
            time.sleep(1)

            self.dismiss_2fa_popup()

        except PlaywrightTimeoutError:
            self.logger.error("[ERROR] Login failed - timeout or incorrect credentials")
            raise Exception("Login failed - please check credentials and network connection")
        except Exception as e:
            self.logger.error(f"[ERROR] Login process failed: {e}")
            raise Exception(f"Login failed: {e}")

    def navigate_and_export_csv_for_date_range(self, start_date, end_date, range_label,
                                               end_operator="<=", retry_count=0, max_retries=2):
        """Navigate to filters and export CSV for a specific date range with retry logic."""
        try:
            self.logger.info(f"\n[EXPORT] Processing date range: {range_label} ({start_date} to {end_date if end_date else 'now'})")
            if retry_count > 0:
                self.logger.info(f"[RETRY] Attempt {retry_count + 1} of {max_retries + 1}")

            self.page.goto(self.config.jira_filter_url, timeout=self.config.default_timeout)
            time.sleep(5)

            # Build JQL query with date range
            if end_date and end_operator:
                jql_query = f'{self.config.base_jql} AND created >= "{start_date}" AND created {end_operator} "{end_date}" ORDER BY created DESC'
            else:
                # No end date - get everything from start date onwards
                jql_query = f'{self.config.base_jql} AND created >= "{start_date}" ORDER BY created DESC'

            self.logger.info(f"[FILTER] Applying JQL: {jql_query}")

            # Switch to JQL mode and apply filter
            try:
                jql_button = self.page.locator('[data-testid="issue-navigator.ui.refinement-bar.search-mode-switcher.toggle-button.advanced"]')
                jql_button.wait_for(state="visible", timeout=10000)
                is_jql_active = jql_button.get_attribute('aria-checked')
                if is_jql_active == 'false':
                    self.logger.info("[FILTER] Switching to JQL mode...")
                    jql_button.click()
                    time.sleep(3)
                    self.logger.info("[SUCCESS] Switched to JQL mode")
                else:
                    self.logger.info("[INFO] Already in JQL mode")
            except Exception as e:
                self.logger.warning(f"[WARNING] Could not verify/switch JQL mode: {e}")

            time.sleep(2)

            # Enter JQL query
            try:
                jql_input = self.page.get_by_test_id("jql-editor-input")
                jql_input.wait_for(state="visible", timeout=10000)
                jql_input.click()
                time.sleep(1)

                self.page.keyboard.press("Control+A")
                time.sleep(0.5)
                self.page.keyboard.press("Delete")
                time.sleep(0.5)

                jql_input.fill(jql_query)
                time.sleep(1)
                jql_input.press("Enter")
                self.logger.info("[SUCCESS] JQL filter applied via input field")

            except Exception as e1:
                self.logger.warning(f"[WARNING] Primary JQL input failed: {e1}, trying keyboard typing...")
                try:
                    self.page.keyboard.press("Control+A")
                    time.sleep(0.5)
                    self.page.keyboard.type(jql_query)
                    time.sleep(1)
                    self.page.keyboard.press("Enter")
                    self.logger.info("[SUCCESS] JQL filter applied via keyboard")
                except Exception as e2:
                    self.logger.error(f"[ERROR] Both JQL input methods failed: {e2}")
                    raise Exception("Could not apply JQL filter")

            time.sleep(5)
            self.logger.info("[SUCCESS] Waiting for JQL filter results to load")

            # Export CSV
            self.logger.info(f"[EXPORT] Starting CSV export for {range_label}")
            self._cleanup_partial_downloads()

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
            expected_filename = f"Jira_{range_label}_{timestamp}.csv"
            expected_path = Path(self.config.download_path) / expected_filename

            self.logger.info(f"[EXPORT] Expected file: {expected_filename}")

            download_dir = Path(self.config.download_path)
            existing_files = set(download_dir.glob("*.csv")) if download_dir.exists() else set()

            # Use longer timeout for open-ended last range
            export_timeout = self.config.export_timeout
            if end_date is None:
                export_timeout = export_timeout + 60000  # Add extra minute for open-ended range
                self.logger.info(f"[EXPORT] Using extended timeout of {export_timeout/1000:.0f} seconds for open-ended range")

            # Open Export menu
            more_actions_btn = self.page.locator(
                "xpath=/html/body/div[1]/div[2]/div[2]/div/div/div/div/div[1]/div/div[1]/div/div[2]/div/div/div[4]/button"
            )
            more_actions_btn.wait_for(state="attached", timeout=15000)
            more_actions_btn.wait_for(state="visible", timeout=15000)
            more_actions_btn.click(force=True)
            self.page.wait_for_timeout(1000)

            export_menu = self.page.locator(
                'button[data-testid="issue-navigator-action-export-issues.ui.filter-button--trigger"]'
            )
            export_menu.wait_for(state="visible", timeout=15000)
            export_menu.click(force=True)
            self.page.wait_for_timeout(800)

            excel_export = self.page.locator(
                'div[data-testid="issue-navigator-action-export-issues.ui.span"]:has-text("Excel CSV - all fields")'
            )
            excel_export.wait_for(state="visible", timeout=15000)
            with self.page.expect_download(timeout=export_timeout) as download_info:
                excel_export.click(force=True)
            download = download_info.value
            self.logger.info(f"[EXPORT] Download started: {download.suggested_filename}")

            # Optional: wait for confirmation text
            try:
                self.logger.info("[EXPORT] Waiting for export completion confirmation...")
                self.page.locator("text=Export complete").first.wait_for(timeout=60000)
                self.logger.info("[EXPORT] Export completion confirmed by Jira")
                time.sleep(2)
            except PlaywrightTimeoutError:
                self.logger.warning("[EXPORT] Export completion text not found, proceeding")
            except Exception as e:
                self.logger.warning(f"[EXPORT] Error waiting for completion confirmation: {e}")

            # Save the downloaded file as expected filename
            download.save_as(str(expected_path))

            actual_file_path = self._verify_and_locate_csv(expected_path, existing_files)

            if actual_file_path:
                file_size = actual_file_path.stat().st_size
                self.logger.info(f"[SUCCESS] CSV for {range_label} exported successfully")
                self.logger.info(f"[FILE] Saved to: {actual_file_path.name}")
                self.logger.info(f"[SIZE] File size: {file_size:,} bytes")

                if self._validate_csv_content(actual_file_path):
                    return str(actual_file_path.absolute())
                else:
                    self.logger.error(f"[ERROR] Downloaded CSV for {range_label} appears to be invalid")
                    return None
            else:
                raise Exception(f"Downloaded file for {range_label} not found or invalid")

        except PlaywrightTimeoutError as timeout_error:
            self.logger.error(f"[TIMEOUT] Export timeout for {range_label}: {timeout_error}")

            # Retry logic for timeouts
            if retry_count < max_retries:
                self.logger.info(f"[RETRY] Retrying export for {range_label}...")
                time.sleep(5)  # Wait before retry
                return self.navigate_and_export_csv_for_date_range(
                    start_date, end_date, range_label, end_operator,
                    retry_count + 1, max_retries
                )
            else:
                self.logger.error(f"[ERROR] Max retries reached for {range_label}")
                return None

        except Exception as e:
            self.logger.error(f"[ERROR] Export failed for {range_label}: {e}")

            # Retry logic for general errors
            if retry_count < max_retries:
                self.logger.info(f"[RETRY] Retrying export for {range_label}...")
                time.sleep(5)
                return self.navigate_and_export_csv_for_date_range(
                    start_date, end_date, range_label, end_operator,
                    retry_count + 1, max_retries
                )
            else:
                return None

    def download_all_date_ranges(self):
        """Download CSVs for all configured date ranges. Fails if ANY range fails."""
        try:
            self.logger.info("\n" + "=" * 70)
            self.logger.info("[MULTI-RANGE] Starting multi-range download process")
            self.logger.info("=" * 70)

            self._setup_browser_context()
            self.login_to_jira()

            downloaded_paths = []
            failed_ranges = []

            for idx, date_range in enumerate(self.config.date_ranges, 1):
                self.logger.info(f"\n[RANGE {idx}/{len(self.config.date_ranges)}] Processing {date_range['label']}")

                csv_path = self.navigate_and_export_csv_for_date_range(
                    date_range['start'],
                    date_range.get('end'),
                    date_range['label'],
                    date_range.get('end_operator', '<=')
                )

                if csv_path:
                    downloaded_paths.append(csv_path)
                    self.logger.info(f"[SUCCESS] Range {idx} completed: {date_range['label']}")
                else:
                    failed_ranges.append(date_range['label'])
                    self.logger.error(f"[ERROR] Failed to download {date_range['label']}")
                    # Don't continue if a range fails - fail immediately
                    self.cleanup_browser()
                    return False, failed_ranges

                if idx < len(self.config.date_ranges):
                    time.sleep(3)

            self.cleanup_browser()

            # All ranges must succeed
            if len(downloaded_paths) != len(self.config.date_ranges):
                self.logger.error(f"[ERROR] Only {len(downloaded_paths)}/{len(self.config.date_ranges)} ranges downloaded")
                return False, failed_ranges

            self.downloaded_csv_paths = downloaded_paths
            self.logger.info(f"\n[SUCCESS] Downloaded all {len(downloaded_paths)} date ranges successfully")

            return True, []

        except Exception as e:
            self.logger.error(f"[ERROR] Multi-range download failed: {e}")
            self.cleanup_browser()
            return False, ["Unknown - Exception occurred"]

    def combine_csv_files(self):
        """Combine multiple CSV files into a single DataFrame."""
        try:
            self.logger.info("\n[COMBINE] Combining downloaded CSV files...")

            if not self.downloaded_csv_paths:
                raise Exception("No CSV files to combine")

            dataframes = []

            for idx, csv_path in enumerate(self.downloaded_csv_paths, 1):
                self.logger.info(f"[LOAD {idx}/{len(self.downloaded_csv_paths)}] Loading: {Path(csv_path).name}")

                try:
                    df = pd.read_csv(csv_path, encoding='utf-8', low_memory=False)
                except UnicodeDecodeError:
                    self.logger.info("[ENCODING] UTF-8 failed, trying Latin-1...")
                    try:
                        df = pd.read_csv(csv_path, encoding='latin-1', low_memory=False)
                    except UnicodeDecodeError:
                        self.logger.info("[ENCODING] Latin-1 failed, trying CP1252...")
                        df = pd.read_csv(csv_path, encoding='cp1252', low_memory=False)

                self.logger.info(f"  • Loaded {len(df):,} rows")
                dataframes.append(df)

            combined_df = pd.concat(dataframes, ignore_index=True)

            if 'Issue key' in combined_df.columns:
                before_dedup = len(combined_df)
                combined_df = combined_df.drop_duplicates(subset=['Issue key'], keep='first')
                after_dedup = len(combined_df)
                duplicates_removed = before_dedup - after_dedup

                if duplicates_removed > 0:
                    self.logger.info(f"[DEDUP] Removed {duplicates_removed} duplicate records")

            self.logger.info(f"[SUCCESS] Combined dataset: {len(combined_df):,} total rows")

            return combined_df

        except Exception as e:
            self.logger.error(f"[ERROR] Failed to combine CSV files: {e}")
            raise

    def _cleanup_partial_downloads(self):
        """Clean up any partial or temporary download files."""
        try:
            download_dir = Path(self.config.download_path)
            if not download_dir.exists():
                return

            patterns = ["*.tmp", "*.part", "*.crdownload", "*Unconfirmed*"]
            cleaned_count = 0

            for pattern in patterns:
                for file_path in download_dir.glob(pattern):
                    try:
                        file_path.unlink()
                        cleaned_count += 1
                        self.logger.debug(f"[CLEANUP] Removed partial download: {file_path.name}")
                    except Exception:
                        pass

            if cleaned_count > 0:
                self.logger.info(f"[CLEANUP] Removed {cleaned_count} partial download file(s)")

        except Exception as e:
            self.logger.warning(f"[WARNING] Cleanup warning: {e}")

    def _verify_and_locate_csv(self, expected_path, existing_files):
        """Verify download and locate the actual CSV file using multiple methods."""
        try:
            download_dir = Path(self.config.download_path)

            if expected_path.exists() and expected_path.stat().st_size > 0:
                self.logger.info(f"[VERIFY] Found file at expected path: {expected_path.name}")
                return expected_path

            current_files = set(download_dir.glob("*.csv")) if download_dir.exists() else set()
            new_files = current_files - existing_files

            if new_files:
                latest_file = max(new_files, key=lambda f: f.stat().st_mtime)
                self.logger.info(f"[VERIFY] Found new CSV file: {latest_file.name}")
                return latest_file

            five_minutes_ago = time.time() - 300
            jira_patterns = ["*jira*", "*Jirareporthelpdesk*", "*helpdesk*", "*issues*"]

            for pattern in jira_patterns:
                for file_path in download_dir.glob(pattern):
                    if (file_path.suffix.lower() == '.csv' and
                        file_path.stat().st_mtime > five_minutes_ago and
                        file_path.stat().st_size > 0):
                        self.logger.info(f"[VERIFY] Found recent Jira CSV: {file_path.name}")
                        return file_path

            csv_files = list(download_dir.glob("*.csv"))
            if csv_files:
                latest_csv = max(csv_files, key=lambda f: f.stat().st_mtime)
                if latest_csv.stat().st_mtime > five_minutes_ago:
                    self.logger.warning(f"[VERIFY] Using most recent CSV: {latest_csv.name}")
                    return latest_csv

            self.logger.error("[ERROR] No valid CSV file found")
            return None

        except Exception as e:
            self.logger.error(f"[ERROR] File verification failed: {e}")
            return None

    def _validate_csv_content(self, file_path: Path):
        """Validate that the CSV file contains expected Jira data."""
        try:
            if file_path.stat().st_size < 100:
                self.logger.error("[VALIDATE] CSV file too small")
                return False

            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                first_line = f.readline().strip()

            expected_headers = ['Issue key', 'Summary', 'Status', 'Created', 'Assignee']
            header_matches = sum(1 for header in expected_headers if header.lower() in first_line.lower())

            if header_matches >= 3:
                self.logger.info(f"[VALIDATE] CSV structure valid ({header_matches}/5 expected headers found)")
                return True
            else:
                self.logger.error(f"[VALIDATE] CSV structure invalid ({header_matches}/5 expected headers found)")
                self.logger.error(f"[VALIDATE] First line: {first_line[:200]}...")
                return False

        except Exception as e:
            self.logger.error(f"[VALIDATE] CSV validation failed: {e}")
            return False

    def cleanup_browser(self):
        """Clean up browser resources."""
        try:
            if self.page:
                self.page.close()
            if self.context:
                self.context.close()
            if self.browser:
                self.browser.close()
            self.logger.info("[CLEANUP] Browser cleanup completed")
        except Exception as e:
            self.logger.warning(f"[WARNING] Browser cleanup warning: {e}")

    # ========================= EMAIL VIA OUTLOOK (MAPI) =========================

    def _send_email_with_outlook(self, subject, html_body, to_list, cc_list=None, bcc_list=None,
                                 attachments=None, inline_images=None, display=False):
        """
        Send an email via Outlook (MAPI). Supports HTML, inline images (cid), and attachments.

        inline_images: dict[str -> bytes or Path]
            Keys are content IDs (e.g., "table_image").
            Values can be raw PNG bytes or a Path to an image file.
        """
        if not WIN32_AVAILABLE:
            self.logger.error("[EMAIL] win32com is not available; cannot send via Outlook.")
            return False

        try:
            import tempfile
            from pathlib import Path as _Path

            outlook = win32.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)  # olMailItem

            # Recipients
            mail.Subject = subject
            mail.To = ";".join(to_list or [])
            if cc_list:
                mail.CC = ";".join(cc_list)
            if bcc_list:
                mail.BCC = ";".join(bcc_list)

            # HTML body
            mail.HTMLBody = html_body

            # Inline images
            tmp_files = []
            if inline_images:
                for cid, content in inline_images.items():
                    # Write bytes to temp file if needed
                    if isinstance(content, (bytes, bytearray)):
                        tmp_path = _Path(tempfile.gettempdir()) / f"{cid}.png"
                        tmp_path.write_bytes(content)
                        tmp_files.append(tmp_path)
                        attach_path = str(tmp_path)
                    else:
                        attach_path = str(_Path(content))

                    attachment = mail.Attachments.Add(attach_path)
                    # Set Content-ID property so <img src="cid:..."> works
                    pa = attachment.PropertyAccessor
                    pa.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)

            # Add file attachments
            if attachments:
                for att in attachments:
                    mail.Attachments.Add(str(att))

            if display:
                mail.Display()  # preview before sending
            else:
                mail.Send()

            self.logger.info("[EMAIL] Email sent successfully via Outlook")
            return True

        except Exception as e:
            self.logger.error(f"[EMAIL] Outlook send failed: {e}")
            return False

    def send_failure_notification_email(self, error_message):
        """Send failure notification via Outlook."""
        try:
            self.logger.info("[EMAIL] Sending failure notification (Outlook)")

            current_date = datetime.now()
            date_str = current_date.strftime("%d-%m-%Y %H:%M:%S")

            body_html = f"""
            <html><body>
            <div style='font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto;'>
                <div style='background-color: #dc3545; color: white; padding: 15px; border-radius: 5px 5px 0 0;'>
                    <h2 style='margin: 0;'>JIRA Export Automation Failed</h2>
                </div>
                <div style='padding: 20px; border: 1px solid #ddd; border-top: none; border-radius: 0 0 5px 5px;'>
                    <p style='color: #444; font-size: 16px;'>Hello,</p>
                    <p style='color: #444; font-size: 16px;'>The automated JIRA CSV export process failed on <b>{date_str} IST</b>.</p>
                    <p style='color: #721c24; background-color: #f8d7da; padding: 10px; border-radius: 5px;'>
                        <strong>Error Details:</strong><br>
                        {error_message}
                    </p>
                    <p style='color: #444;'>Please check the system and retry manually if needed.</p>
                </div>
            </div>
            </body></html>
            """

            subject = f"JIRA Export Failed - {current_date.strftime('%d-%m-%Y %H%M%S')}"
            cfg = self.config.failure_email_config

            return self._send_email_with_outlook(
                subject=subject,
                html_body=body_html,
                to_list=cfg.get('toaddr', []),
                cc_list=cfg.get('cc', []),
                bcc_list=cfg.get('bcc', []),
                attachments=None,
                inline_images=None,
                display=False
            )

        except Exception as e:
            self.logger.error(f"[ERROR] Failed to send failure notification (Outlook): {e}")
            return False

    # ========================= CSV PROCESSING METHODS =========================

    def filter_and_process_data(self, df):
        """Filter dataframe and do all data processing in pandas for speed."""
        original_count = len(df)
        self.logger.info(f"[PROCESS] Starting data processing on {original_count:,} records...")

        # Step 1: Remove cancelled tickets
        cancelled_removed = 0
        if 'Status' in df.columns:
            cancelled_patterns = ['cancelled', 'canceled', 'cancel']
            mask = ~df['Status'].astype(str).str.lower().str.contains('|'.join(cancelled_patterns), na=False)
            df = df[mask].copy()
            cancelled_removed = original_count - len(df)
            if cancelled_removed > 0:
                self.logger.info(f"[FILTER] Removed {cancelled_removed} cancelled tickets")

        # Step 2: Fill blank resolved dates with updated dates
        blank_assignee_count = 0
        if 'Resolved' in df.columns and 'Updated' in df.columns:
            blank_resolved_mask = (
                df['Resolved'].isna()
                | (df['Resolved'].astype(str).str.strip() == '')
                | (df['Resolved'].astype(str).str.lower().isin(['nan', 'none', 'null']))
            )
            blank_resolved_count = blank_resolved_mask.sum()
            if blank_resolved_count > 0:
                df.loc[blank_resolved_mask, 'Resolved'] = df.loc[blank_resolved_mask, 'Updated']
                self.logger.info(f"[PROCESS] Filled {blank_resolved_count} blank resolved dates with updated dates")

        # Step 3: Remove records with blank/unassigned assignees
        if 'Assignee' in df.columns:
            blank_assignee_mask = (
                df['Assignee'].isna()
                | (df['Assignee'].astype(str).str.strip() == '')
                | (df['Assignee'].astype(str).str.lower().isin(['nan', 'none', 'null', 'unassigned']))
            )
            blank_assignee_count = blank_assignee_mask.sum()
            if blank_assignee_count > 0:
                df = df[~blank_assignee_mask].copy()
                self.logger.info(f"[PROCESS] Removed {blank_assignee_count} records with blank/unassigned assignees")

        self.logger.info(f"[PROCESS] Data processing complete:")
        self.logger.info(f"  • Original records: {original_count:,}")
        if cancelled_removed > 0:
            self.logger.info(f"  • Cancelled tickets removed: {cancelled_removed:,}")
        if blank_assignee_count > 0:
            self.logger.info(f"  • Unassigned tickets removed: {blank_assignee_count:,}")
        self.logger.info(f"  • Final dataset: {len(df):,} records")

        return df

    def validate_data_structure(self, df):
        """Validate data structure and log findings."""
        self.logger.info("[VALIDATE] Validating data structure...")

        required_columns = ['Summary', 'Issue key', 'Issue id', 'Status', 'Created', 'Resolved', 'Assignee']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            self.logger.warning(f"[WARNING] Missing expected columns: {missing_columns}")
        else:
            self.logger.info("[SUCCESS] All key columns found")

        if 'Created' in df.columns:
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    parsed_dates = pd.to_datetime(df['Created'], errors='coerce', infer_datetime_format=True)
                    valid_dates = parsed_dates.notna().sum()
                    self.logger.info(f"[INFO] Valid 'Created' dates: {valid_dates}/{len(df)}")
            except Exception as e:
                self.logger.warning(f"[WARNING] Could not validate Created dates: {e}")

        if 'Resolved' in df.columns:
            try:
                non_empty_resolved = df[df['Resolved'].astype(str).str.strip() != '']['Resolved'].count()
                self.logger.info(f"[INFO] Records with resolved dates: {non_empty_resolved}/{len(df)}")
            except Exception as e:
                self.logger.warning(f"[WARNING] Could not validate Resolved dates: {e}")

        return len(missing_columns) == 0

    def create_working_copy(self):
        """Create a working copy of the template file."""
        if not os.path.exists(self.config.template_path):
            raise FileNotFoundError(f"Template file not found: {self.config.template_path}")

        current_date = datetime.now()
        date_str = current_date.strftime("%d-%m-%Y")

        output_filename = f"Jira report {date_str}.xlsx"
        output_path = os.path.join(self.config.output_folder, output_filename)

        shutil.copy2(self.config.template_path, output_path)
        self.logger.info(f"[COPY] Created working copy: {os.path.basename(output_path)}")

        return output_path

    # ========================= EXCEL FORMATTING METHODS =========================

    def apply_excel_formatting(self, worksheet, dataframe):
        """Apply readable formatting to Excel worksheet with reasonable performance."""
        header_fill = PatternFill(start_color="80C4E8", end_color="80C4E8", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        max_row = len(dataframe) + 1
        max_col = len(dataframe.columns)

        # Header styling only (fast)
        for col in range(1, max_col + 1):
            c = worksheet.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align
            c.border = thin_border

        # Row heights
        for row in range(1, max_row + 1):
            worksheet.row_dimensions[row].height = 16

        # Column widths (approximate)
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def format_date_columns(self, worksheet, dataframe):
        """Apply Excel number formats to columns that are datetime-like."""
        date_column_patterns = {"created", "updated", "resolved", "due date", "last viewed"}

        # 1) Convert in DataFrame first (much faster/reliable)
        for col in dataframe.columns:
            if col.strip().lower() in date_column_patterns:
                try:
                    dataframe[col] = pd.to_datetime(dataframe[col], errors='coerce')
                except Exception:
                    pass

        # 2) Re-write date columns into the sheet with number format
        for col_idx, col_name in enumerate(dataframe.columns, 1):
            col_lower = col_name.strip().lower()
            if col_lower in date_column_patterns:
                self.logger.info(f"[FORMAT] Applying date format to column: {col_name}")
                for row_idx in range(2, len(dataframe) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    val = dataframe.iloc[row_idx - 2, col_idx - 1]
                    if pd.notna(val):
                        try:
                            ts = pd.to_datetime(val).to_pydatetime()
                            cell.value = ts
                            cell.number_format = 'M/D/YYYY H:MM AM/PM'
                        except Exception:
                            pass

    def update_data_sheet_with_formatting(self, workbook_path, df, sheet_name="Jira Ticket Details"):
        """Update the data sheet and apply formatting (data already processed in pandas)."""
        workbook = openpyxl.load_workbook(workbook_path)

        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {workbook.sheetnames}")

        worksheet = workbook[sheet_name]
        worksheet.delete_rows(1, worksheet.max_row)

        # Ensure likely date columns are datetime before writing
        candidate_date_cols = {'created', 'updated', 'resolved', 'due date', 'last viewed'}
        for col in df.columns:
            if col.strip().lower() in candidate_date_cols:
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                except Exception:
                    pass

        self.logger.info(f"[EXCEL] Writing {len(df):,} rows and {len(df.columns)} columns to '{sheet_name}' sheet")
        for row_data in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row_data)

        self.logger.info("[EXCEL] Applying formatting...")
        self.apply_excel_formatting(worksheet, df)
        self.format_date_columns(worksheet, df)

        workbook.save(workbook_path)
        workbook.close()

        self.logger.info(f"[SUCCESS] Data sheet '{sheet_name}' updated and formatted successfully")
        return True

    # ========================= EXCEL (COM) DYNAMIC HEADERS & PIVOT =========================

    def update_dynamic_headers(self, file_path, df):
        """Update pivot table headers with current date range."""
        if not WIN32_AVAILABLE:
            self.logger.warning("[WARNING] win32com not available - headers not updated")
            return False

        excel_app = None
        try:
            self.logger.info("[EXCEL] Updating dynamic headers...")

            start_date_str = "1st January 2025"

            current_date = datetime.now()
            day = current_date.day
            # Handle 11th-13th correctly
            if 11 <= day <= 13:
                suffix = "th"
            elif day % 10 == 1:
                suffix = "st"
            elif day % 10 == 2:
                suffix = "nd"
            elif day % 10 == 3:
                suffix = "rd"
            else:
                suffix = "th"
            end_date_str = f"{day}{suffix} {current_date.strftime('%B %Y')}"

            self.logger.info(f"[HEADER] Will show: {start_date_str} till {end_date_str}")

            excel_app = win32.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            workbook = excel_app.Workbooks.Open(os.path.abspath(file_path))

            updated_headers = 0

            # Progress Tracker
            try:
                progress_sheet = workbook.Worksheets("Progress Tracker")
                for row in range(1, 6):
                    for col in range(1, 16):
                        try:
                            cell = progress_sheet.Cells(row, col)
                            cell_value = str(cell.Value) if cell.Value else ""
                            if "User vs Resolved case from" in cell_value and "till" in cell_value:
                                new_header = f"User vs Resolved case from {start_date_str} till {end_date_str}"
                                cell.Value = new_header
                                self.logger.info(f"[HEADER] Updated Progress Tracker header in {chr(64+col)}{row}")
                                updated_headers += 1
                        except Exception:
                            continue
            except Exception as e:
                self.logger.warning(f"[WARNING] Could not update Progress Tracker headers: {e}")

            # Issue vs Tickets
            try:
                issues_sheet = workbook.Worksheets("Issue vs Tickets")
                for row in range(1, 6):
                    for col in range(1, 13):
                        try:
                            cell = issues_sheet.Cells(row, col)
                            cell_value = str(cell.Value) if cell.Value else ""
                            if "Issue vs Tickets from" in cell_value and "till" in cell_value:
                                new_header = f"Issue vs Tickets from {start_date_str} till {end_date_str}"
                                cell.Value = new_header
                                self.logger.info(f"[HEADER] Updated Issue vs Tickets header in {chr(64+col)}{row}")
                                updated_headers += 1
                        except Exception:
                            continue
            except Exception as e:
                self.logger.warning(f"[WARNING] Could not update Issue vs Tickets headers: {e}")

            # Pending header updates
            try:
                progress_sheet = workbook.Worksheets("Progress Tracker")
                current_month = current_date.strftime('%B')
                current_year = current_date.year

                for row in range(1, 6):
                    for col in range(1, 16):
                        try:
                            cell = progress_sheet.Cells(row, col)
                            cell_value = str(cell.Value) if cell.Value else ""
                            if "User vs Pending for" in cell_value:
                                new_header = f"User vs Pending for {current_month} {current_year}"
                                cell.Value = new_header
                                self.logger.info(f"[HEADER] Updated Pending header in {chr(64+col)}{row}")
                                updated_headers += 1
                        except Exception:
                            continue
            except Exception as e:
                self.logger.warning(f"[WARNING] Could not update Pending headers: {e}")

            workbook.Save()
            workbook.Close(False)

            if updated_headers > 0:
                self.logger.info(f"[SUCCESS] Successfully updated {updated_headers} header(s)")
                return True
            else:
                self.logger.warning("[WARNING] No headers were updated - check if headers contain expected text")
                return False

        except Exception as e:
            self.logger.error(f"[ERROR] Error updating dynamic headers: {str(e)}")
            return False
        finally:
            try:
                if excel_app:
                    excel_app.Quit()
            except Exception:
                pass

    def refresh_pivot_tables_excel(self, file_path):
        """Refresh all pivot tables using Excel COM automation."""
        if not WIN32_AVAILABLE:
            self.logger.warning("[WARNING] win32com not available - pivot tables not refreshed")
            return False

        excel_app = None
        try:
            self.logger.info("[EXCEL] Opening Excel to refresh pivot tables...")

            excel_app = win32.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            workbook = excel_app.Workbooks.Open(os.path.abspath(file_path))

            pivot_count = 0
            for worksheet in workbook.Worksheets:
                try:
                    for pivot_table in worksheet.PivotTables():
                        try:
                            pivot_table.PivotCache().Refresh()
                            pivot_count += 1
                            self.logger.info(f"[REFRESH] Refreshed pivot table '{pivot_table.Name}' in sheet '{worksheet.Name}'")
                        except Exception as e:
                            self.logger.error(f"[ERROR] Error refreshing pivot table '{pivot_table.Name}': {str(e)}")
                except Exception:
                    # Some sheets may not have PivotTables() enumerable
                    continue

            workbook.Save()
            workbook.Close(False)

            if pivot_count > 0:
                self.logger.info(f"[SUCCESS] Successfully refreshed {pivot_count} pivot table(s)")
                return True
            else:
                self.logger.warning("[WARNING] No pivot tables found to refresh")
                return False

        except Exception as e:
            self.logger.error(f"[ERROR] Error during pivot table refresh: {str(e)}")
            self.logger.error("[INFO] Make sure Microsoft Excel is installed and the file is not open elsewhere")
            return False
        finally:
            try:
                if excel_app:
                    excel_app.Quit()
            except Exception:
                pass

    def extract_pivot_table_data(self, file_path, sheet_name="Progress Tracker"):
        """Extract pivot table data for email."""
        if not WIN32_AVAILABLE:
            self.logger.warning("[WARNING] win32com not available - using fallback method for pivot data")
            return None

        excel_app = None
        try:
            self.logger.info(f"[EXTRACT] Extracting PivotTable1 data from '{sheet_name}' sheet...")

            excel_app = win32.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            workbook = excel_app.Workbooks.Open(os.path.abspath(file_path))
            worksheet = workbook.Worksheets(sheet_name)

            pivot_table = None
            try:
                for pt in worksheet.PivotTables():
                    if pt.Name == "PivotTable1":
                        pivot_table = pt
                        break
            except Exception:
                pass

            if pivot_table is None:
                self.logger.info("[INFO] PivotTable1 not found, using first available pivot table")
                pivot_table = worksheet.PivotTables(1)

            pt_range = pivot_table.TableRange1

            data = []
            for row in range(1, pt_range.Rows.Count + 1):
                row_data = []
                for col in range(1, pt_range.Columns.Count + 1):
                    cell_value = pt_range.Cells(row, col).Value
                    row_data.append(cell_value if cell_value is not None else "")
                data.append(row_data)

            workbook.Close(False)

            if not data or len(data) < 3:
                return None

            df = pd.DataFrame(data)

            # Heuristic header reconstruction
            if len(df) >= 3:
                headers = df.iloc[1].tolist()
                if not headers or all(str(h).strip() == "" for h in headers):
                    return None

                if headers[0] in ("", None) or (isinstance(headers[0], float) and pd.isna(headers[0])):
                    headers[0] = "Name"

                df.columns = headers
                df = df.iloc[3:].reset_index(drop=True)

                # Coerce numeric columns
                name_col = df.columns[0]
                numeric_cols = [c for c in df.columns if c != name_col]
                for col in numeric_cols:
                    try:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
                    except Exception:
                        pass

                # Keep rows with any non-zero numeric or a proper name
                if numeric_cols:
                    keep_mask = (
                        df[numeric_cols].sum(axis=1) > 0
                    ) | (
                        ~df[name_col].astype(str).str.strip().str.lower().isin(['name', '', 'nan'])
                    )
                    df = df[keep_mask]

                # Drop fully empty columns
                df = df.loc[:, (df != '').any(axis=0)]

            self.logger.info(f"[SUCCESS] Extracted PivotTable1 with {len(df)} rows and {len(df.columns)} columns")
            return df

        except Exception as e:
            self.logger.error(f"[ERROR] Error extracting PivotTable1: {str(e)}")
            return None
        finally:
            try:
                if excel_app:
                    excel_app.Quit()
            except Exception:
                pass

    def create_table_image(self, pivot_data):
        """Create a professional, readable table image from pivot data."""
        if pivot_data is None or pivot_data.empty:
            return None

        try:
            num_rows = len(pivot_data) + 1
            num_cols = len(pivot_data.columns)

            fig_width = max(16, num_cols * 1.2)
            fig_height = max(8, num_rows * 0.6 + 2)

            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            fig.patch.set_facecolor('white')
            ax.axis('off')

            table_data = [list(pivot_data.columns)]
            for _, row in pivot_data.iterrows():
                formatted_row = []
                for i, val in enumerate(row.values):
                    if i == 0:
                        formatted_row.append(str(val))
                    else:
                        try:
                            num_val = int(float(val))
                            formatted_row.append(f"{num_val:,}" if num_val >= 1000 else str(num_val))
                        except Exception:
                            formatted_row.append(str(val))
                table_data.append(formatted_row)

            # Column widths
            col_widths = []
            for i, header in enumerate(table_data[0]):
                if i == 0:
                    max_length = max(len(str(header)), max(len(str(r[i])) for r in table_data[1:]))
                    col_widths.append(min(0.25, max(0.15, max_length * 0.012)))
                else:
                    col_widths.append(0.08)

            table = ax.table(
                cellText=table_data[1:],
                colLabels=table_data[0],
                cellLoc='center',
                loc='center',
                colWidths=col_widths
            )

            table.auto_set_font_size(False)
            table.set_fontsize(11)
            table.scale(1, 2.5)

            # Header styling
            for i in range(len(table_data[0])):
                cell = table[(0, i)]
                cell.set_facecolor("#ff0606")
                cell.set_text_props(weight='bold', color='white', fontsize=12)
                cell.set_height(0.08)
                cell.get_text().set_horizontalalignment('center')

            # Body styling
            last_idx = len(table_data) - 1
            for i in range(1, len(table_data)):
                for j in range(len(table_data[0])):
                    cell = table[(i, j)]
                    if i == last_idx:
                        cell.set_facecolor("#9DA1A4")
                        cell.set_text_props(weight='bold', fontsize=12)
                    elif i % 2 == 0:
                        cell.set_facecolor("#f7f7f7")
                    else:
                        cell.set_facecolor('white')
                    cell.set_height(0.06)
                    cell.set_text_props(fontsize=12)

                    if j == 0:
                        cell.get_text().set_horizontalalignment('left')
                        cell.get_text().set_position((0.05, 0.5))
                    else:
                        cell.get_text().set_horizontalalignment('right')
                        cell.get_text().set_position((0.95, 0.5))

            for _, cell in table.get_celld().items():
                cell.set_linewidth(0.5)
                cell.set_edgecolor("#686464")

            plt.tight_layout(pad=1.0)

            img_buffer = io.BytesIO()
            plt.savefig(
                img_buffer,
                format='png',
                dpi=200,
                bbox_inches='tight',
                facecolor='white',
                edgecolor='none',
                pad_inches=0.3
            )
            img_buffer.seek(0)
            plt.close(fig)

            self.logger.info("[SUCCESS] Table image created successfully")
            return img_buffer.getvalue()

        except Exception as e:
            self.logger.error(f"[ERROR] Failed to create table image: {e}")
            return None

    def send_success_email(self, file_path, pivot_data):
        """Send success email via Outlook with inline table image and Excel attachment."""
        try:
            self.logger.info("[EMAIL] Preparing success email (Outlook)")

            current_date = datetime.now()
            date_DD = current_date.strftime("%d-%m-%Y")
            time_T = current_date.strftime("%H:%M:%S")
            time_string = current_date.strftime("%H%M%S")

            # Inline table image (if pivot data is present)
            inline_images = {}
            html_table_block = "<p style='color: #dc3545; text-align: center;'>Table could not be generated - please check the attached Excel file.</p>"
            if pivot_data is not None and not pivot_data.empty:
                table_img_bytes = self.create_table_image(pivot_data)
                if table_img_bytes:
                    inline_images["table_image"] = table_img_bytes
                    html_table_block = (
                        '<div style="text-align: center; margin: 25px 0;">'
                        '<img src="cid:table_image" style="max-width: 100%; height: auto; border: 2px solid #ddd; '
                        'border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.1);" />'
                        '</div>'
                    )

            # HTML body
            body_html = f"""
            <html>
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
            </head>
            <body>
            <div style='font-family: "Segoe UI", Arial, sans-serif; max-width: 1000px; margin: 0 auto; background-color: #f8f9fa;'>
                <div style='background-color: #1c4e80; color: white; padding: 20px; border-radius: 8px 8px 0 0;'>
                    <h2 style='margin: 0; font-size: 24px; text-align: center;'>JIRA IT-Support Ticket Tracker</h2>
                </div>
                <div style='padding: 30px; background-color: white; border: 1px solid #ddd; border-top: none; border-radius: 0 0 8px 8px;'>
                    <p style='color: #444; font-size: 16px; margin-bottom: 10px;'>Hello Team,</p>
                    <p style='color: #444; font-size: 16px; margin-bottom: 25px;'>
                        Below is the JIRA IT-Support Ticket Tracker for <strong>{date_DD} {time_T} IST</strong>:
                    </p>
                    {html_table_block}
                    <p style='color: #444; margin-top: 35px; margin-bottom: 10px;'>
                        If you have any questions regarding this report, please feel free to reach out.
                    </p>
                    <p style='color: #444; margin-top: 25px; margin-bottom: 0;'>
                        Thanks,<br><strong>Team IT</strong>
                    </p>
                </div>
                <div style='font-size: 12px; color: #777; text-align: center; margin-top: 15px; padding: 15px;'>
                    This is an automated report. Please do not reply to this email.
                </div>
            </div>
            </body>
            </html>
            """

            subject = f"JIRA IT-Support Ticket Tracker {date_DD} {time_string}"
            attachments = [file_path]  # Excel report
            cfg = self.config.email_config

            return self._send_email_with_outlook(
                subject=subject,
                html_body=body_html,
                to_list=cfg.get('toaddr', []),
                cc_list=cfg.get('cc', []),
                bcc_list=cfg.get('bcc', []),
                attachments=attachments,
                inline_images=inline_images,
                display=False  # set True if you want to preview before sending
            )

        except Exception as e:
            self.logger.error(f"[ERROR] Failed to send success email via Outlook: {e}")
            return False

    # ========================= MAIN WORKFLOW METHOD =========================

    def run_complete_automation(self):
        """Run the complete automation workflow with multi-range download."""
        self.logger.info("[START] Starting complete JIRA automation workflow with multi-range download")

        # Optional guard: ensure Outlook available if email enabled
        if self.config.send_email and self.config.email_via_outlook and not WIN32_AVAILABLE:
            self.logger.error("[EMAIL] Outlook (win32com) not available on this machine. Install Outlook or disable email sending.")
            return False

        try:
            # Step 1: Download CSVs for all date ranges
            self.logger.info("\n" + "=" * 70)
            self.logger.info("[STEP 1] DOWNLOADING CSVs FOR MULTIPLE DATE RANGES")
            self.logger.info("=" * 70)

            download_success, failed_ranges = self.download_all_date_ranges()

            if not download_success:
                if failed_ranges:
                    error_msg = f"Failed to download date range(s): {', '.join(failed_ranges)}. All date ranges must download successfully for the report to be generated."
                else:
                    error_msg = "Failed to download CSVs from Jira - check login credentials and network connection"

                self.logger.error(f"[ERROR] {error_msg}")
                if self.config.send_email:
                    self.send_failure_notification_email(error_msg)
                return False

            # Step 2: Combine all CSV files
            self.logger.info("\n" + "=" * 70)
            self.logger.info("[STEP 2] COMBINING CSV FILES INTO SINGLE DATASET")
            self.logger.info("=" * 70)

            df = self.combine_csv_files()

            if df is None or len(df) == 0:
                error_msg = "Failed to combine CSV files or no data found"
                self.logger.error(f"[ERROR] {error_msg}")
                if self.config.send_email:
                    self.send_failure_notification_email(error_msg)
                return False

            # Step 3: Process and filter data (ALL processing in pandas for speed)
            self.logger.info("\n" + "=" * 70)
            self.logger.info("[STEP 3] PROCESSING AND FILTERING DATA")
            self.logger.info("=" * 70)

            self.validate_data_structure(df)
            df = self.filter_and_process_data(df)  # All filtering and processing in pandas

            if len(df) == 0:
                error_msg = "No data found after filtering and processing"
                self.logger.error(f"[ERROR] {error_msg}")
                if self.config.send_email:
                    self.send_failure_notification_email(error_msg)
                return False

            # Step 4: Create Excel report
            self.logger.info("\n" + "=" * 70)
            self.logger.info("[STEP 4] CREATING EXCEL REPORT")
            self.logger.info("=" * 70)

            output_file = self.create_working_copy()
            self.update_data_sheet_with_formatting(output_file, df)

            # Step 5: Update headers and refresh pivot tables
            self.logger.info("\n" + "=" * 70)
            self.logger.info("[STEP 5] UPDATING HEADERS AND REFRESHING PIVOT TABLES")
            self.logger.info("=" * 70)

            header_success = self.update_dynamic_headers(output_file, df)
            pivot_success = self.refresh_pivot_tables_excel(output_file)

            # Step 6: Send success email with report
            if self.config.send_email:
                self.logger.info("\n" + "=" * 70)
                self.logger.info("[STEP 6] SENDING SUCCESS EMAIL WITH REPORT")
                self.logger.info("=" * 70)

                pivot_data = self.extract_pivot_table_data(output_file, self.config.email_sheet)
                email_success = self.send_success_email(output_file, pivot_data)
            else:
                email_success = True
                self.logger.info("\n[STEP 6] Email sending disabled in configuration")

            # Final summary
            self.logger.info("\n" + "=" * 70)
            self.logger.info("[COMPLETE] AUTOMATION WORKFLOW COMPLETED SUCCESSFULLY!")
            self.logger.info("=" * 70)
            self.logger.info(f"Downloaded {len(self.downloaded_csv_paths)} date range files:")
            for path in self.downloaded_csv_paths:
                self.logger.info(f"  • {Path(path).name}")
            self.logger.info(f"Processed {len(df):,} rows of data (cancelled tickets excluded)")
            self.logger.info(f"Generated report: {os.path.basename(output_file)}")
            self.logger.info(f"Location: {os.path.dirname(output_file)}")

            if header_success:
                self.logger.info("Headers updated with current date range")
            else:
                self.logger.warning("Headers may need manual update")

            if pivot_success:
                self.logger.info("Pivot tables refreshed successfully")
            else:
                self.logger.warning("Pivot tables may need manual refresh in Excel")

            if self.config.send_email and email_success:
                all_recipients = list(set(self.config.email_config.get('toaddr', []) +
                                          self.config.email_config.get('cc', []) +
                                          self.config.email_config.get('bcc', [])))
                self.logger.info(f"Success email sent to {len(all_recipients)} recipients")

            return True

        except Exception as e:
            error_msg = f"Automation workflow failed: {str(e)}"
            self.logger.error(f"[ERROR] {error_msg}")

            if self.config.send_email:
                self.send_failure_notification_email(error_msg)

            return False


# ========================= ENTRY POINT =========================
def main():
    """Main function to run the integrated automation system."""
    print("=" * 70)
    print("JIRA MULTI-RANGE EXPORT AND REPORT AUTOMATION SYSTEM")
    print("=" * 70)

    try:
        config = IntegratedJiraConfig()

        print("\n[VALIDATION] Checking system requirements...")

        if not os.path.exists(config.template_path):
            print(f"[ERROR] Template file not found: {config.template_path}")
            print("Please ensure the Excel template exists at the specified location.")
            return 1

        if not WIN32_AVAILABLE:
            print("[WARNING] win32com not available - Outlook emails and Excel COM features will not work")
            print("   Install with: pip install pywin32 (and ensure Outlook & Excel are installed)")

        print(f"\n[CONFIG] Download Path: {config.download_path}")
        print(f"[CONFIG] Template Path: {config.template_path}")
        print(f"[CONFIG] Output Folder: {config.output_folder}")
        print(f"[CONFIG] Jira Base URL: {config.jira_base_url}")
        print(f"[CONFIG] Email Enabled: {config.send_email} (Outlook: {config.email_via_outlook})")

        print("\n[CONFIG] Date ranges to download:")
        for idx, dr in enumerate(config.date_ranges, 1):
            print(f"  {idx}. {dr['label']}: {dr['start']} to {dr['end']}")

        automation = IntegratedJiraAutomation(config)
        success = automation.run_complete_automation()

        if success:
            print("\n[SUCCESS] Complete automation workflow finished successfully!")
            print("Check the logs above for detailed information about the process.")
            return 0
        else:
            print("\n[FAILED] Automation workflow failed")
            print("Check the logs above for error details.")
            return 1

    except KeyboardInterrupt:
        print("\n[INTERRUPT] Process interrupted by user (Ctrl+C)")
        return 1
    except Exception as e:
        print(f"\n[ERROR] Fatal error: {e}")
        import traceback
        print("\nDetailed error information:")
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    try:
        print("[DEPENDENCY] Checking for required modules...")

        required_modules = {
            'playwright': 'playwright',
            'pandas': 'pandas',
            'openpyxl': 'openpyxl',
            'matplotlib': 'matplotlib'
        }

        missing_modules = []
        for module_name, install_name in required_modules.items():
            try:
                __import__(module_name)
                print(f"[OK] {module_name} is available")
            except ImportError:
                missing_modules.append(install_name)
                print(f"[MISSING] {module_name} not found")

        if missing_modules:
            print(f"\n[ERROR] Missing required modules: {', '.join(missing_modules)}")
            print(f"Please install with: pip install {' '.join(missing_modules)}")
            print("Also run: playwright install")
            sys.exit(1)

        if not WIN32_AVAILABLE:
            print("[OPTIONAL] pywin32 is not available - install with: pip install pywin32")
            print("Outlook sending & Excel pivot refresh require Windows with Outlook/Excel installed.")

        print("[DEPENDENCY] All required modules are available")

    except Exception as e:
        print(f"[ERROR] Error checking dependencies: {e}")

    try:
        result = main()
        sys.exit(result)
    except KeyboardInterrupt:
        print("\n[INTERRUPT] Process interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n[FATAL ERROR] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)